import datetime as dt
import io
import logging
from datetime import timedelta

from django.db.models import Count, OuterRef, Q, Subquery
from django.utils import timezone
from openpyxl import Workbook
from weasyprint import HTML

from apps.audit.models import AuditLogEntry
from apps.tasks.models import Task

logger = logging.getLogger(__name__)

# Cap on how many stuck tasks to send to LLM/PDF/Excel; total count is always reported.
STUCK_TASK_SAMPLE = 10


def _coerce_period_bounds(date_from, date_to):
    """Normalize loose period inputs into timezone-aware datetimes.

    Callers pass either date strings ("2026-04-08"), date objects, or already-aware
    datetime objects. Date-only values are expanded to a full day in the project
    timezone:
        - date_from → start-of-day (00:00:00.000)
        - date_to   → end-of-day (23:59:59.999999)

    Without this expansion, "daily" summaries (period_from == period_to) would
    silently return zero audit-log rows because the implicit midnight cutoff
    excludes everything that happened later in the day.
    """
    def _to_dt(value, end_of_day):
        if value is None:
            return None
        if isinstance(value, dt.datetime):
            if timezone.is_naive(value):
                return timezone.make_aware(value, timezone.get_current_timezone())
            return value
        if isinstance(value, dt.date):
            naive = dt.datetime.combine(
                value,
                dt.time.max if end_of_day else dt.time.min,
            )
            return timezone.make_aware(naive, timezone.get_current_timezone())
        # Treat as ISO date string.
        try:
            d = dt.date.fromisoformat(str(value))
        except ValueError:
            # Last-resort: let Django handle whatever the caller passed.
            return value
        naive = dt.datetime.combine(
            d, dt.time.max if end_of_day else dt.time.min,
        )
        return timezone.make_aware(naive, timezone.get_current_timezone())

    return _to_dt(date_from, end_of_day=False), _to_dt(date_to, end_of_day=True)


def _percentile(sorted_values, pct):
    """Linear-interpolation percentile (pct is 0-100). Returns None for empty input."""
    if not sorted_values:
        return None
    if len(sorted_values) == 1:
        return sorted_values[0]
    k = (len(sorted_values) - 1) * (pct / 100.0)
    lo = int(k)
    hi = min(lo + 1, len(sorted_values) - 1)
    if lo == hi:
        return sorted_values[lo]
    return sorted_values[lo] + (sorted_values[hi] - sorted_values[lo]) * (k - lo)


def _compute_lead_times(done_entries):
    """Lead time (task.created_at -> done event timestamp), in hours, per task.

    Returns a list of floats. If a task has multiple "done" events in the
    queryset (shouldn't normally happen, but possible after status flip-flops),
    only the most recent event is used.
    """
    rows = list(
        done_entries
        .values("task_id", "timestamp", "task__created_at")
        .order_by("task_id", "-timestamp")
    )
    seen = set()
    durations = []
    for r in rows:
        tid = r["task_id"]
        if tid in seen:
            continue
        seen.add(tid)
        created_at = r["task__created_at"]
        done_at = r["timestamp"]
        if created_at is None or done_at is None:
            continue
        durations.append((done_at - created_at).total_seconds() / 3600.0)
    return durations


def _compute_cycle_times(done_entries, audit_qs):
    """Cycle time (most recent in_progress event -> done event), in hours, per task.

    Cycle time only exists for tasks that actually passed through "in_progress"
    before being marked done. Tasks that jumped straight to done are skipped.
    """
    done_rows = list(
        done_entries.values("task_id", "timestamp").order_by("task_id", "-timestamp")
    )
    latest_done = {}
    for r in done_rows:
        latest_done.setdefault(r["task_id"], r["timestamp"])
    if not latest_done:
        return []

    in_progress_rows = list(
        audit_qs.filter(
            action=AuditLogEntry.Action.STATUS_CHANGE,
            new_value="in_progress",
            task_id__in=latest_done.keys(),
        )
        .values("task_id", "timestamp")
        .order_by("task_id", "-timestamp")
    )

    latest_in_progress = {}
    for r in in_progress_rows:
        tid = r["task_id"]
        done_ts = latest_done[tid]
        if r["timestamp"] > done_ts:
            continue
        existing = latest_in_progress.get(tid)
        if existing is None or r["timestamp"] > existing:
            latest_in_progress[tid] = r["timestamp"]

    cycles = []
    for tid, done_ts in latest_done.items():
        ip_ts = latest_in_progress.get(tid)
        if ip_ts is not None:
            cycles.append((done_ts - ip_ts).total_seconds() / 3600.0)
    return cycles


def _summarize_durations(durations):
    """Return {avg, median, p90, count} dict in hours, rounded."""
    if not durations:
        return {"avg_hours": None, "median_hours": None, "p90_hours": None, "count": 0}
    sorted_durations = sorted(durations)
    return {
        "avg_hours": round(sum(sorted_durations) / len(sorted_durations), 1),
        "median_hours": round(_percentile(sorted_durations, 50), 1),
        "p90_hours": round(_percentile(sorted_durations, 90), 1),
        "count": len(sorted_durations),
    }


def get_report_data(date_from=None, date_to=None, client_id=None, organization=None):
    """Aggregate task metrics for a period.

    `base_qs` (org-scoped, all-time) is used for current-state metrics: total counts,
    overdue, unassigned, stuck-waiting — these describe "right now."

    `period_qs` (org-scoped, restricted by date_from/date_to via task creation date)
    drives per-engineer/per-client/per-tag breakdowns and `created_in_period`. The
    audit log drives `closed_in_period` and resolution-time metrics so that completions
    are attributed to the period in which the status change happened, not when the
    task was originally created.
    """
    date_from, date_to = _coerce_period_bounds(date_from, date_to)

    base_qs = Task.objects.all()
    if organization:
        base_qs = base_qs.filter(organization=organization)
    if client_id:
        base_qs = base_qs.filter(client_id=client_id)

    total = base_qs.count()

    status_counts = dict(
        base_qs.values("status").annotate(c=Count("id")).values_list("status", "c")
    )
    by_status = {s: status_counts.get(s, 0) for s in Task.Status.values}

    priority_counts = dict(
        base_qs.values("priority").annotate(c=Count("id")).values_list("priority", "c")
    )
    by_priority = {p: priority_counts.get(p, 0) for p in Task.Priority.values}

    now = timezone.now()
    overdue = base_qs.filter(deadline__lt=now).exclude(
        status__in=["done", "archived"]
    ).count()

    # Period-specific metrics.
    period_qs = base_qs
    audit_period_filter = Q()
    if organization:
        audit_period_filter &= Q(task__organization=organization)
    if client_id:
        audit_period_filter &= Q(task__client_id=client_id)

    has_period = bool(date_from or date_to)
    if date_from:
        period_qs = period_qs.filter(created_at__gte=date_from)
        audit_period_filter &= Q(timestamp__gte=date_from)
    if date_to:
        period_qs = period_qs.filter(created_at__lte=date_to)
        audit_period_filter &= Q(timestamp__lte=date_to)

    if has_period:
        created_in_period = period_qs.count()
        done_entries = AuditLogEntry.objects.filter(
            audit_period_filter,
            action=AuditLogEntry.Action.STATUS_CHANGE,
            new_value="done",
        )
        closed_in_period = done_entries.values("task_id").distinct().count()

        # Lead time: created -> done. Computed in Python so we can also get median/p90.
        lead_durations = _compute_lead_times(done_entries)
        lead_time = _summarize_durations(lead_durations)

        # Cycle time: most recent in_progress -> done. Needs the unfiltered audit
        # log scoped to org so we can find in_progress events that may have
        # happened before the period window.
        org_audit_qs = AuditLogEntry.objects.all()
        if organization:
            org_audit_qs = org_audit_qs.filter(task__organization=organization)
        if client_id:
            org_audit_qs = org_audit_qs.filter(task__client_id=client_id)
        cycle_durations = _compute_cycle_times(done_entries, org_audit_qs)
        cycle_time = _summarize_durations(cycle_durations)

        avg_resolution_time_hours = lead_time["avg_hours"]
        completion_rate = (
            round(closed_in_period / created_in_period * 100, 1)
            if created_in_period > 0
            else None
        )
    else:
        created_in_period = total
        closed_in_period = base_qs.filter(status="done").count()
        avg_resolution_time_hours = None
        completion_rate = None
        lead_time = {"avg_hours": None, "median_hours": None, "p90_hours": None, "count": 0}
        cycle_time = {"avg_hours": None, "median_hours": None, "p90_hours": None, "count": 0}

    unassigned_count = (
        base_qs.filter(assignees__isnull=True)
        .exclude(status__in=["done", "archived"])
        .count()
    )

    # Deadline proximity: active tasks with deadline in next 48 hours.
    deadline_soon_cutoff = now + timedelta(hours=48)
    approaching_deadline_qs = (
        base_qs.filter(deadline__gt=now, deadline__lte=deadline_soon_cutoff)
        .exclude(status__in=["done", "archived"])
        .order_by("deadline")
    )
    approaching_deadline = [
        {
            "id": t.id,
            "title": t.title,
            "priority": t.priority,
            "deadline": t.deadline.isoformat(),
            "hours_remaining": round((t.deadline - now).total_seconds() / 3600.0, 1),
        }
        for t in approaching_deadline_qs[:10]
    ]

    # New overdue vs inherited: tasks whose deadline crossed "now" during the period.
    if has_period and date_from:
        new_overdue = (
            base_qs.filter(deadline__gte=date_from, deadline__lt=now)
            .exclude(status__in=["done", "archived"])
            .count()
        )
    else:
        new_overdue = None
    inherited_overdue = (overdue - new_overdue) if new_overdue is not None else None

    # Status transitions in period: count audit log STATUS_CHANGE entries by flow.
    if has_period:
        transition_rows = list(
            AuditLogEntry.objects.filter(
                audit_period_filter,
                action=AuditLogEntry.Action.STATUS_CHANGE,
            )
            .values("old_value", "new_value")
            .annotate(count=Count("id"))
            .order_by("-count")
        )
        status_transitions = [
            {
                "from": r["old_value"],
                "to": r["new_value"],
                "count": r["count"],
            }
            for r in transition_rows
        ]
    else:
        status_transitions = []

    # Stuck-waiting tasks (current-state, not period-scoped).
    waiting_tasks = base_qs.filter(status="waiting")
    three_days_ago = now - timedelta(days=3)
    latest_waiting_entry = (
        AuditLogEntry.objects.filter(
            task=OuterRef("pk"),
            action=AuditLogEntry.Action.STATUS_CHANGE,
            new_value="waiting",
        )
        .order_by("-timestamp")
        .values("timestamp")[:1]
    )
    stuck_qs = (
        waiting_tasks.annotate(waiting_since=Subquery(latest_waiting_entry))
        .filter(waiting_since__lt=three_days_ago)
        .order_by("waiting_since")
    )
    stuck_total = stuck_qs.count()
    stuck_sample = []
    for t in stuck_qs[:STUCK_TASK_SAMPLE]:
        waiting_hours = round((now - t.waiting_since).total_seconds() / 3600.0, 1) if t.waiting_since else None
        stuck_sample.append({
            "id": t.id,
            "title": t.title,
            "priority": t.priority,
            "waiting_hours": waiting_hours,
        })
    stuck_waiting = {
        "count": stuck_total,
        "sample": stuck_sample,
    }

    # Period-scoped breakdowns. The previous implementation used base_qs (all-time),
    # which silently leaked lifetime totals into weekly/daily summaries.
    breakdown_qs = period_qs if has_period else base_qs

    by_client_qs = (
        breakdown_qs.filter(client__isnull=False)
        .values("client__id", "client__name")
        .annotate(
            total=Count("id", distinct=True),
            done=Count("id", filter=Q(status="done"), distinct=True),
        )
        .order_by("-total")
    )

    by_engineer_qs = (
        breakdown_qs.filter(assignees__isnull=False)
        .values("assignees__id", "assignees__first_name", "assignees__last_name")
        .annotate(
            assigned=Count("id", distinct=True),
            done=Count("id", filter=Q(status="done"), distinct=True),
        )
        .order_by("-assigned")
    )

    by_tag_qs = (
        breakdown_qs.filter(tags__isnull=False)
        .values("tags__id", "tags__name")
        .annotate(count=Count("id", distinct=True))
        .order_by("-count")
    )

    return {
        "period": {
            "from": date_from.date().isoformat() if date_from else None,
            "to": date_to.date().isoformat() if date_to else None,
        },
        "tasks": {
            "total": total,
            "by_status": by_status,
            "by_priority": by_priority,
            "created_in_period": created_in_period,
            "closed_in_period": closed_in_period,
            "overdue": overdue,
            "avg_resolution_time_hours": avg_resolution_time_hours,
            "unassigned_count": unassigned_count,
            "stuck_waiting": stuck_waiting,
            "completion_rate": completion_rate,
            "lead_time": lead_time,
            "cycle_time": cycle_time,
            "approaching_deadline": approaching_deadline,
            "overdue_new": new_overdue,
            "overdue_inherited": inherited_overdue,
            "status_transitions": status_transitions,
        },
        "by_client": [
            {
                "client_id": c["client__id"],
                "client_name": c["client__name"],
                "total": c["total"],
                "done": c["done"],
            }
            for c in by_client_qs
        ],
        "by_engineer": [
            {
                "engineer_id": e["assignees__id"],
                "engineer_name": f"{e['assignees__first_name']} {e['assignees__last_name']}".strip(),
                "assigned": e["assigned"],
                "done": e["done"],
            }
            for e in by_engineer_qs
        ],
        "by_tag": [
            {
                "tag_id": t["tags__id"],
                "tag_name": t["tags__name"],
                "count": t["count"],
            }
            for t in by_tag_qs
        ],
    }


def generate_pdf_report(data):
    logger.info("Generating PDF report period=%s", data.get("period"))
    tasks = data["tasks"]
    avg_res = tasks.get("avg_resolution_time_hours")
    avg_res_display = f"{avg_res} hours" if avg_res is not None else "N/A"
    comp_rate = tasks.get("completion_rate")
    comp_rate_display = f"{comp_rate}%" if comp_rate is not None else "N/A"
    stuck = tasks.get("stuck_waiting", {})
    stuck_count = stuck.get("count", 0)

    tag_rows = "".join(
        f"<tr><td>{t['tag_name']}</td><td>{t['count']}</td></tr>"
        for t in data.get("by_tag", [])
    )

    html_content = f"""
    <html>
    <head><style>
        body {{ font-family: sans-serif; padding: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin: 16px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f5f5f5; }}
        h1 {{ color: #1976d2; }}
    </style></head>
    <body>
        <h1>Task Management Report</h1>
        <p>Period: {data['period']['from'] or 'All'} to {data['period']['to'] or 'All'}</p>

        <h2>Summary</h2>
        <table>
            <tr><th>Metric</th><th>Value</th></tr>
            <tr><td>Total Tasks</td><td>{tasks['total']}</td></tr>
            <tr><td>Created in Period</td><td>{tasks['created_in_period']}</td></tr>
            <tr><td>Closed in Period</td><td>{tasks['closed_in_period']}</td></tr>
            <tr><td>Overdue</td><td>{tasks['overdue']}</td></tr>
            <tr><td>Avg Resolution Time</td><td>{avg_res_display}</td></tr>
            <tr><td>Completion Rate</td><td>{comp_rate_display}</td></tr>
            <tr><td>Unassigned Active Tasks</td><td>{tasks.get('unassigned_count', 0)}</td></tr>
            <tr><td>Stuck Waiting (3+ days)</td><td>{stuck_count}</td></tr>
        </table>

        <h2>By Status</h2>
        <table>
            <tr><th>Status</th><th>Count</th></tr>
            {''.join(f"<tr><td>{k}</td><td>{v}</td></tr>" for k, v in tasks['by_status'].items())}
        </table>

        <h2>By Priority</h2>
        <table>
            <tr><th>Priority</th><th>Count</th></tr>
            {''.join(f"<tr><td>{k}</td><td>{v}</td></tr>" for k, v in tasks['by_priority'].items())}
        </table>

        <h2>By Client</h2>
        <table>
            <tr><th>Client</th><th>Total</th><th>Done</th></tr>
            {''.join(f"<tr><td>{c['client_name']}</td><td>{c['total']}</td><td>{c['done']}</td></tr>" for c in data['by_client'])}
        </table>

        <h2>By Engineer</h2>
        <table>
            <tr><th>Engineer</th><th>Assigned</th><th>Done</th></tr>
            {''.join(f"<tr><td>{e['engineer_name']}</td><td>{e['assigned']}</td><td>{e['done']}</td></tr>" for e in data['by_engineer'])}
        </table>

        {"<h2>By Tag</h2><table><tr><th>Tag</th><th>Count</th></tr>" + tag_rows + "</table>" if tag_rows else ""}
    </body>
    </html>
    """
    pdf_buffer = io.BytesIO()
    HTML(string=html_content).write_pdf(pdf_buffer)
    pdf_buffer.seek(0)
    return pdf_buffer


def generate_excel_report(data):
    logger.info("Generating Excel report period=%s", data.get("period"))
    wb = Workbook()
    tasks = data["tasks"]

    avg_res = tasks.get("avg_resolution_time_hours")
    comp_rate = tasks.get("completion_rate")
    stuck = tasks.get("stuck_waiting", {})

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    ws_summary.append(["Total Tasks", tasks["total"]])
    ws_summary.append(["Created in Period", tasks["created_in_period"]])
    ws_summary.append(["Closed in Period", tasks["closed_in_period"]])
    ws_summary.append(["Overdue", tasks["overdue"]])
    ws_summary.append(["Avg Resolution Time (hours)", avg_res if avg_res is not None else "N/A"])
    ws_summary.append(["Completion Rate (%)", comp_rate if comp_rate is not None else "N/A"])
    ws_summary.append(["Unassigned Active Tasks", tasks.get("unassigned_count", 0)])
    ws_summary.append(["Stuck Waiting (3+ days)", stuck.get("count", 0)])
    ws_summary.append([])
    ws_summary.append(["Status", "Count"])
    for k, v in tasks["by_status"].items():
        ws_summary.append([k, v])
    ws_summary.append([])
    ws_summary.append(["Priority", "Count"])
    for k, v in tasks["by_priority"].items():
        ws_summary.append([k, v])

    ws_clients = wb.create_sheet("By Client")
    ws_clients.append(["Client", "Total", "Done"])
    for c in data["by_client"]:
        ws_clients.append([c["client_name"], c["total"], c["done"]])

    ws_engineers = wb.create_sheet("By Engineer")
    ws_engineers.append(["Engineer", "Assigned", "Done"])
    for e in data["by_engineer"]:
        ws_engineers.append([e["engineer_name"], e["assigned"], e["done"]])

    by_tag = data.get("by_tag", [])
    if by_tag:
        ws_tags = wb.create_sheet("By Tag")
        ws_tags.append(["Tag", "Count"])
        for t in by_tag:
            ws_tags.append([t["tag_name"], t["count"]])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
